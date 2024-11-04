import fitz
from pathlib import Path
import os
from openpyxl import Workbook
from openpyxl.styles import Border, Side, Alignment


def create_workbook(properties):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = 'Query'

    # Add New Data Formula
    properties.append(['Jumlah', '', f'=sum(C2:C{len(properties)})'])

    # Fill Data
    total_data = fill_workbook(properties, sheet)

    # Apply Style
    style_workbook(sheet, total_data)

    filename = input("Input File Name: ")
    workbook.save(f'{filename}.xlsx')
    workbook.close()


def fill_workbook(data, sheet):
    for row in data:
        sheet.append(row)
    return len(data)


def auto_fit_column_widths(sheet):
    column_widths = {}

    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None:
                value_length = len(str(cell.value))
                column = cell.column_letter

                if column in column_widths:
                    column_widths[column] = max(
                        column_widths[column], value_length)
                else:
                    column_widths[column] = value_length

    for column, width in column_widths.items():
        sheet.column_dimensions[column].width = width + 2


def style_workbook(sheet, total_data):
    # Preparation Border
    thin_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )

    center_alignment = Alignment(horizontal='center', vertical='center')

    # Apply Border
    for row in sheet.iter_rows(min_row=1, max_row=total_data, min_col=1, max_col=3):
        for cell in row:
            if cell.row != 1 and cell.column != 2:
                cell.border = thin_border
                cell.alignment = center_alignment
            else:
                cell.border = thin_border

    auto_fit_column_widths(sheet)


def pdf_checking(pdf_file, properties, i):
    pdf_document = fitz.open(pdf_file)
    properties.append([i+1, pdf_file.stem, pdf_document.page_count])
    print(f"Files Name: {pdf_file.stem}\nTotal Pages: {
          pdf_document.page_count}\n----------------------------------------")
    return pdf_document.page_count


def pdf_folder(folder_path):
    pdf_files = []
    for pdf_path in Path(folder_path).rglob('*.pdf'):
        pdf_files.append(pdf_path)
    return pdf_files


def pdf_data(pdf_files, pdf_checking, pdf_properties):
    pdf_total = 0
    for i, pdf_file in enumerate(pdf_files):
        pdf_total += pdf_checking(pdf_file, pdf_properties, i)
    print(f"Actual Total: {pdf_total} Pages\nTotal Files: {
          len(pdf_files)} Files")


def main():
    folder_path = os.getcwd()

    pdf_path = pdf_folder(folder_path)
    pdf_properties = []
    pdf_properties.append(['No', 'Name', 'Pages'])

    pdf_data(pdf_path, pdf_checking, pdf_properties)

    check_input = input("Extract data to excel? (Y/N): ").strip().lower()

    if (check_input == 'y' or check_input == 'yes'):
        create_workbook(pdf_properties)
    else:
        print('Proccess Completed.')


if __name__ == '__main__':
    main()
